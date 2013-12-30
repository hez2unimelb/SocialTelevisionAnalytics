'''
Created on 9 Dec, 2013

@author: zhaohe
'''
import xlrd
import openpyxl
import datetime
import os
class ExcelReader(object):
    '''
    classdocs
    '''


    def __init__(self, fileName):
        '''
        Constructor
        '''
        self.workbook = openpyxl.reader.excel.load_workbook(fileName)
        
    def printNames(self, location):
        workbook = xlrd.open_workbook(location)
        print workbook.sheet_names()
    def printEachRow(self,location,sheetName):
        workbook = xlrd.open_workbook(location)
        worksheet = workbook.sheet_by_name(sheetName)
        num_rows = worksheet.nrows - 1
        curr_row = -1
        while curr_row < num_rows:
            curr_row += 1
            row = worksheet.row(curr_row)
            print row
    def getLocationTweets(self,sheetName):
        worksheet = self.workbook.get_sheet_by_name(sheetName)
        rows= worksheet.range(worksheet.calculate_dimension())
        data=[[]]
        for i in range(len(rows)) :
            vector=[]
            for j in range(len(rows[i])):
                cell =worksheet.cell(row=i,column=j)
                if cell.get_coordinate()[0]=='D':
                    vector.append(cell.value)
                if cell.get_coordinate()[0]=='E':
                    vector.append(cell.value)
                if len(vector)==2:
                    data.append(vector)
                    break
            del vector
        del data[0]
        return data
    def getDayComment(self,sheetName):
        worksheet = self.workbook.get_sheet_by_name(sheetName)
        rows= worksheet.range(worksheet.calculate_dimension())
        data=[[]]
        for i in range(len(rows)) :
            vector=[]
            for j in range(len(rows[i])):
                cell =worksheet.cell(row=i,column=j)
                if cell.get_coordinate()[0]=='A':
                    vector.append(str(cell.value.date()))
                if cell.get_coordinate()[0]=='B':
                    vector.append(cell.value)
                if len(vector)==2:
                    data.append(vector)
                    break
            del vector
        del data[0]
        return data
    def getHourTweets(self,sheetName):
        worksheet = self.workbook.get_sheet_by_name(sheetName)
        rows= worksheet.range(worksheet.calculate_dimension())
        data=[[]]
        for i in range(len(rows)) :
            vector=[]
            for j in range(len(rows[i])):
                cell =worksheet.cell(row=i,column=j)
                if cell.get_coordinate()[0]=='A':
                    vector.append(str(cell.value))
                if cell.get_coordinate()[0]=='B':
                    vector.append(cell.value)
                if len(vector)==2:
                    data.append(vector)
                    break
            del vector
        del data[0]
        return data
    def getTriple(self,sheetName):
        worksheet = self.workbook.get_sheet_by_name(sheetName)
        rows= worksheet.range(worksheet.calculate_dimension())
        data=[[]]
        for i in range(len(rows)) :
            vector=[]
            for j in range(len(rows[i])):
                cell =worksheet.cell(row=i,column=j)
                if cell.get_coordinate()[0]=='A':
                    vector.append(str(cell.value))
                if cell.get_coordinate()[0]=='B':
                    vector.append(cell.value)
                if cell.get_coordinate()[0]=='C':
                    vector.append(cell.value)
                    print cell.value
                if cell.get_coordinate()[0]=='D':
                    vector.append(cell.value)
                if len(vector)==4:
                    data.append(vector)
                    break
            del vector
        del data[0]
        del data[0]
        return data
    def getOneDayHourTweets(self,sheetName):
        worksheet = self.workbook.get_sheet_by_name(sheetName)
        rows= worksheet.range(worksheet.calculate_dimension())
        data=[[]]
        for i in range(len(rows)) :
            vector=[]
            for j in range(len(rows[i])):
                cell =worksheet.cell(row=i,column=j)
                if cell.get_coordinate()[0]=='A':
                    vector.append(str(cell.value))
                if cell.get_coordinate()[0]=='B':
                    vector.append(cell.value)
                if len(vector)==2:
                    data.append(vector)
                    break
            del vector
        del data[0]
        del data[0]
        return data  
    def writeCSV(self,csvName,columnName,data):
       
        fileHandle=open('D:/academic/workspace/SocialTV/static/csvfile/'+csvName,'w')
        strColumnName=''
        for name in columnName:
            strColumnName+=name+','
        strColumnName=strColumnName[:-1]
        fileHandle.write(strColumnName+'\n')
        strData=''
        for d in data:
            for c in d:
                strData+=str(c)+','
            strData=strData[:-1]
            strData+='\n'
        fileHandle.write(strData);              
                
    
        
            
            
            
            
            